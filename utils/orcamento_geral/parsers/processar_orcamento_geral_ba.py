from __future__ import annotations

import argparse
from pathlib import Path
import sys

import pandas as pd
import pyarrow.parquet as pq
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[3]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from project_paths import ORCAMENTO_GERAL_PROCESSADA_DIR, cli_default
from utils.convenios.unificador import build_parquet_table, normalize_preview
from utils.common import STANDARD_COLUMNS
from utils.orcamento_geral.paths import add_scope_argument, default_output_name, uf_raw_dir


ORIGEM_ORCAMENTO_GERAL = "ESTADO_ORCAMENTO_GERAL"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Processa os arquivos de despesas gerais da BA para parquet no schema padrao."
    )
    add_scope_argument(parser)
    parser.add_argument(
        "--input",
        help="Arquivo bruto da BA. Se omitido, usa despesas_*.xlsx do escopo escolhido.",
    )
    parser.add_argument(
        "--output-dir",
        default=cli_default(ORCAMENTO_GERAL_PROCESSADA_DIR),
        help="Pasta de saida para os parquets da trilha de orcamento geral.",
    )
    return parser.parse_args()


def default_input_path(scope: str) -> Path:
    return uf_raw_dir("BA", scope) / "pagamentos_osc_candidatas_cruzadas.csv"


def default_input_dir(scope: str) -> Path:
    return uf_raw_dir("BA", scope)


def read_csv_with_fallback(path: Path) -> pd.DataFrame:
    last_error: Exception | None = None
    for encoding in ("utf-8", "utf-8-sig", "latin1"):
        try:
            return pd.read_csv(path, dtype=str, encoding=encoding, low_memory=False)
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"Falha ao ler {path}") from last_error


def read_source(path: Path) -> pd.DataFrame:
    if path.suffix.lower() in {".xlsx", ".xls"}:
        return pd.DataFrame(iter_general_expense_rows(path))
    return read_csv_with_fallback(path)


def read_default_sources(input_dir: Path) -> pd.DataFrame:
    payment_paths = sorted(input_dir.glob("pagamentos_painel_*_detalhamento.csv"))
    if payment_paths:
        frames = [read_csv_with_fallback(path).assign(_arquivo_origem=path.name) for path in payment_paths]
        return pd.concat(frames, ignore_index=True, sort=False)

    paths = sorted(input_dir.glob("despesas_*.xlsx"))
    if not paths:
        return read_source(input_dir / "pagamentos_osc_candidatas_cruzadas.csv")

    rows: list[dict[str, object]] = []
    for path in paths:
        rows.extend(iter_general_expense_rows(path))
    return pd.DataFrame(rows)


def clean_value(value: object) -> object:
    if value is None:
        return pd.NA
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return pd.NA
    return text


def first_non_empty_value(*values: object) -> object:
    for value in values:
        cleaned = clean_value(value)
        if not pd.isna(cleaned):
            return cleaned
    return pd.NA


def iter_general_expense_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    columns = [str(value).strip() if value is not None else "" for value in header]
    index_by_name = {name: idx for idx, name in enumerate(columns) if name}
    rows: list[dict[str, object]] = []

    def get(values: tuple[object, ...], column: str) -> object:
        index = index_by_name.get(column)
        return values[index] if index is not None and index < len(values) else None

    for values in worksheet.iter_rows(min_row=2, values_only=True):
        rows.append(
            {
                "uf": "BA",
                "origem": ORIGEM_ORCAMENTO_GERAL,
                "ano": first_non_empty_value(get(values, "ANO_EXERCICIO")),
                "valor_total": first_non_empty_value(
                    get(values, "VAL_LIQUIDADO_TOTAL"),
                    get(values, "VAL_EMPENHADO_TOTAL"),
                    get(values, "VAL_PAGO"),
                ),
                "cnpj": pd.NA,
                "nome_osc": first_non_empty_value(get(values, "NOM_UNIDADE_GESTORA"), get(values, "NOM_ORGAO_ORCAMENTO")),
                "mes": first_non_empty_value(get(values, "MES_EXERCICIO")),
                "cod_municipio": pd.NA,
                "municipio": pd.NA,
                "objeto": first_non_empty_value(get(values, "NOM_ACAO_PROGRAMA_GOVERNO"), get(values, "NOM_PROGRAMA_GOVERNO")),
                "modalidade": first_non_empty_value(
                    get(values, "NOM_MODALIDADE_APLICACAO_ORCAMENTO"),
                    get(values, "NOM_ELEMENTO_DESPESA_ORCAMENTO"),
                    get(values, "NOM_GRUPO_DESPESA_ORCAMENTO"),
                ),
                "data_inicio": first_non_empty_value(get(values, "DATA_COMPLETA")),
                "data_fim": pd.NA,
            }
        )

    workbook.close()
    return rows


def first_non_empty(*series: pd.Series | None) -> pd.Series:
    if not series:
        return pd.Series(dtype="string")

    result: pd.Series | None = None
    for current in series:
        if current is None:
            continue
        cleaned = current.astype("string").str.strip().replace("", pd.NA)
        result = cleaned if result is None else result.combine_first(cleaned)

    if result is None:
        return pd.Series(dtype="string")
    return result


def build_ba_budget_frame(source_df: pd.DataFrame) -> pd.DataFrame:
    if set(STANDARD_COLUMNS).issubset(source_df.columns):
        return source_df[STANDARD_COLUMNS]

    if "ANO_EXERCICIO" in source_df.columns:
        return build_ba_general_expense_frame(source_df)

    if {"Recebedor", "CPF/CNPJ", "Valor do Pagamento", "Data do Pagamento"}.issubset(source_df.columns):
        return build_ba_payment_detail_frame(source_df)

    modalidade = first_non_empty(
        source_df.get("tipo_parceria_convenio"),
        source_df.get("tipo_instrumento_convenio"),
        source_df.get("tipo_despesa_convenio"),
    )
    objeto = first_non_empty(source_df.get("objeto_convenio"))
    municipio = first_non_empty(source_df.get("municipio_convenio"))

    mapped = pd.DataFrame(
        {
            "uf": "BA",
            "origem": ORIGEM_ORCAMENTO_GERAL,
            "ano": source_df.get("ano"),
            "valor_total": source_df.get("valor_total"),
            "cnpj": source_df.get("cnpj"),
            "nome_osc": source_df.get("recebedor"),
            "mes": source_df.get("mes"),
            "municipio": municipio,
            "objeto": objeto,
            "modalidade": modalidade,
            "data_inicio": source_df.get("data_inicio_convenio"),
            "data_fim": source_df.get("data_fim_convenio"),
        }
    )

    # A compilacao da BA nao traz um codigo de municipio confiavel.
    mapped["cod_municipio"] = pd.NA

    for column in STANDARD_COLUMNS:
        if column not in mapped.columns:
            mapped[column] = pd.NA
    return mapped[STANDARD_COLUMNS]


def build_ba_payment_detail_frame(source_df: pd.DataFrame) -> pd.DataFrame:
    data_pagamento = pd.to_datetime(source_df.get("Data do Pagamento"), errors="coerce", format="mixed")
    ano = pd.Series(data_pagamento.dt.year, index=source_df.index, dtype="Int64").astype("string")
    mes = pd.Series(data_pagamento.dt.month, index=source_df.index, dtype="Int64").astype("string")

    mapped = pd.DataFrame(
        {
            "uf": "BA",
            "origem": ORIGEM_ORCAMENTO_GERAL,
            "ano": ano,
            "valor_total": source_df.get("Valor do Pagamento"),
            "cnpj": source_df.get("CPF/CNPJ"),
            "nome_osc": source_df.get("Recebedor"),
            "mes": mes,
            "cod_municipio": pd.NA,
            "municipio": pd.NA,
            "objeto": first_non_empty(source_df.get("Unidade Orçamentária"), source_df.get("Órgão")),
            "modalidade": first_non_empty(source_df.get("Nº do Empenho\r"), source_df.get("Nº do Empenho")),
            "data_inicio": source_df.get("Data do Pagamento"),
            "data_fim": pd.NA,
        }
    )

    for column in STANDARD_COLUMNS:
        if column not in mapped.columns:
            mapped[column] = pd.NA
    return mapped[STANDARD_COLUMNS]


def build_ba_general_expense_frame(source_df: pd.DataFrame) -> pd.DataFrame:
    mapped = pd.DataFrame(
        {
            "uf": "BA",
            "origem": ORIGEM_ORCAMENTO_GERAL,
            "ano": source_df.get("ANO_EXERCICIO"),
            "valor_total": first_non_empty(
                source_df.get("VAL_LIQUIDADO_TOTAL"),
                source_df.get("VAL_EMPENHADO_TOTAL"),
                source_df.get("VAL_PAGO"),
            ),
            "cnpj": pd.NA,
            "nome_osc": first_non_empty(source_df.get("NOM_UNIDADE_GESTORA"), source_df.get("NOM_ORGAO_ORCAMENTO")),
            "mes": source_df.get("MES_EXERCICIO"),
            "cod_municipio": pd.NA,
            "municipio": pd.NA,
            "objeto": first_non_empty(source_df.get("NOM_ACAO_PROGRAMA_GOVERNO"), source_df.get("NOM_PROGRAMA_GOVERNO")),
            "modalidade": first_non_empty(
                source_df.get("NOM_MODALIDADE_APLICACAO_ORCAMENTO"),
                source_df.get("NOM_ELEMENTO_DESPESA_ORCAMENTO"),
                source_df.get("NOM_GRUPO_DESPESA_ORCAMENTO"),
            ),
            "data_inicio": source_df.get("DATA_COMPLETA"),
            "data_fim": pd.NA,
        }
    )

    for column in STANDARD_COLUMNS:
        if column not in mapped.columns:
            mapped[column] = pd.NA
    return mapped[STANDARD_COLUMNS]


def main() -> None:
    args = parse_args()
    input_path = Path(args.input) if args.input else None
    input_dir = default_input_dir(args.scope)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    source_df = read_source(input_path) if input_path else read_default_sources(input_dir)
    mapped = build_ba_budget_frame(source_df)
    normalized = normalize_preview(mapped, "BA", require_cnpj=True)

    output_path = output_dir / default_output_name("BA", args.scope)
    pq.write_table(build_parquet_table(normalized), output_path, compression="snappy")

    print(f"Entrada: {input_path or input_dir / 'pagamentos_painel_*_detalhamento.csv'}")
    print(f"Saida: {output_path}")
    print(f"Linhas origem: {len(source_df)}")
    print(f"Linhas parquet: {len(normalized)}")
    print(f"Origem aplicada: {ORIGEM_ORCAMENTO_GERAL}")


if __name__ == "__main__":
    main()

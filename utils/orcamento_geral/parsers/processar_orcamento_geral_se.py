from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
import sys

import pandas as pd
import pyarrow.parquet as pq
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[3]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from project_paths import ORCAMENTO_GERAL_PROCESSADA_DIR, cli_default
from utils.common import STANDARD_COLUMNS
from utils.convenios.unificador import build_parquet_table, normalize_preview
from utils.orcamento_geral.paths import add_scope_argument, default_output_name, uf_raw_dir


ORIGEM_ORCAMENTO_GERAL = "ESTADO_ORCAMENTO_GERAL"
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Processa todos os empenhos de SE como despesas gerais do orcamento estadual."
    )
    add_scope_argument(parser)
    parser.add_argument(
        "--input-dir",
        help="Diretorio com os arquivos empenhos_*.xlsx. Se omitido, usa o caminho padrao do escopo.",
    )
    parser.add_argument(
        "--output-dir",
        default=cli_default(ORCAMENTO_GERAL_PROCESSADA_DIR),
        help="Pasta de saida para os parquets da trilha de orcamento geral.",
    )
    return parser.parse_args()


def default_input_dir(scope: str) -> Path:
    return uf_raw_dir("SE", scope)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return ""
    return " ".join(text.split())


def normalize_document(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = clean_text(value)
    digits = re.sub(r"\D", "", text)
    if len(digits) == 13:
        digits = f"0{digits}"
    return digits


def first_non_empty(*values: object) -> object:
    for value in values:
        text = clean_text(value)
        if text:
            return text
    return pd.NA


def iter_source_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    columns = [clean_text(value) for value in header]
    index_by_name = {name: idx for idx, name in enumerate(columns) if name}
    rows: list[dict[str, object]] = []

    for values in worksheet.iter_rows(min_row=2, values_only=True):
        nome_osc = clean_text(values[index_by_name["nmRazaoSocialPessoa"]]) if "nmRazaoSocialPessoa" in index_by_name else ""
        cnpj = normalize_document(values[index_by_name["nuDocumento"]]) if "nuDocumento" in index_by_name else ""

        row = {
            "uf": "SE",
            "origem": ORIGEM_ORCAMENTO_GERAL,
            "ano": first_non_empty(
                values[index_by_name["dtAnoExercicioCtb"]] if "dtAnoExercicioCtb" in index_by_name else None,
                values[index_by_name["_ano"]] if "_ano" in index_by_name else None,
            ),
            "valor_total": first_non_empty(
                values[index_by_name["vlTotalLiquidadoEmpenho"]] if "vlTotalLiquidadoEmpenho" in index_by_name else None,
                values[index_by_name["vlOriginalEmpenho"]] if "vlOriginalEmpenho" in index_by_name else None,
                values[index_by_name["vlSolicEmpenho"]] if "vlSolicEmpenho" in index_by_name else None,
                values[index_by_name["vlTotalPagoEmpenho"]] if "vlTotalPagoEmpenho" in index_by_name else None,
            ),
            "cnpj": cnpj,
            "nome_osc": nome_osc,
            "mes": first_non_empty(values[index_by_name["_mes"]] if "_mes" in index_by_name else None),
            "cod_municipio": pd.NA,
            "municipio": pd.NA,
            "objeto": first_non_empty(values[index_by_name["dsObjetoLicitacao"]] if "dsObjetoLicitacao" in index_by_name else None),
            "modalidade": first_non_empty(values[index_by_name["nmModalidadeLicitacao"]] if "nmModalidadeLicitacao" in index_by_name else None),
            "data_inicio": first_non_empty(
                values[index_by_name["dtEmissaoEmpenho"]] if "dtEmissaoEmpenho" in index_by_name else None,
                values[index_by_name["dtLancamentoEmpenho"]] if "dtLancamentoEmpenho" in index_by_name else None,
                values[index_by_name["dtGeracaoEmpenho"]] if "dtGeracaoEmpenho" in index_by_name else None,
            ),
            "data_fim": pd.NA,
        }
        rows.append(row)

    workbook.close()
    return rows


def iter_api_rows(path: Path) -> list[dict[str, object]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise ValueError(f"Formato inesperado em {path}: {type(payload).__name__}")

    rows: list[dict[str, object]] = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        nome_osc = clean_text(item.get("nomePessoa"))
        cnpj = normalize_document(item.get("codigoFavorecido"))

        data_empenho = clean_text(item.get("dataEmpenho"))[:10] or pd.NA
        data_parseada = pd.to_datetime(data_empenho, dayfirst=True, errors="coerce") if data_empenho is not pd.NA else pd.NaT
        ano = data_parseada.year if not pd.isna(data_parseada) else first_non_empty(item.get("ano"))
        rows.append(
            {
                "uf": "SE",
                "origem": ORIGEM_ORCAMENTO_GERAL,
                "ano": ano,
                "valor_total": first_non_empty(item.get("valorExecutado"), item.get("valorOriginal")),
                "cnpj": cnpj,
                "nome_osc": nome_osc,
                "mes": data_parseada.month if not pd.isna(data_parseada) else pd.NA,
                "cod_municipio": pd.NA,
                "municipio": pd.NA,
                "objeto": first_non_empty(item.get("descricaoSolicitacao"), item.get("elementoDespesa")),
                "modalidade": first_non_empty(item.get("modalidade"), item.get("elementoDespesa")),
                "data_inicio": data_empenho,
                "data_fim": pd.NA,
            }
        )
    return rows


def build_se_budget_frame(input_dir: Path) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for path in sorted(input_dir.glob("empenhos_*.xlsx")):
        rows.extend(iter_source_rows(path))
    for path in sorted(input_dir.glob("se_empenhos_api_*.json")):
        rows.extend(iter_api_rows(path))
    frame = pd.DataFrame(rows)
    if frame.empty:
        return pd.DataFrame(columns=STANDARD_COLUMNS).astype("string")
    for column in STANDARD_COLUMNS:
        if column not in frame.columns:
            frame[column] = pd.NA
    return frame[STANDARD_COLUMNS]


def main() -> None:
    args = parse_args()
    input_dir = Path(args.input_dir) if args.input_dir else default_input_dir(args.scope)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    mapped = build_se_budget_frame(input_dir)
    normalized = normalize_preview(mapped, "SE", require_cnpj=True)

    output_path = output_dir / default_output_name("SE", args.scope)
    pq.write_table(build_parquet_table(normalized), output_path, compression="snappy")

    print(f"Entrada: {input_dir}")
    print(f"Saida: {output_path}")
    print(f"Linhas parquet: {len(normalized)}")
    print(f"Origem aplicada: {ORIGEM_ORCAMENTO_GERAL}")


if __name__ == "__main__":
    main()
